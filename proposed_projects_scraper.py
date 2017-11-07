from bs4 import BeautifulSoup
import requests
import time
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
import os
import config


#the array to pass to excel
excel_links = []

#url to start
#try:

my_username = str(config.user_id)
my_password = str(config.password)

url = "https://www.dot14.state.pa.us/ECMS/"
driver = webdriver.Chrome()#r'C:\Windows\chromedriver.exe')  # Optional argument, if not specified will search path.
driver.get(url);
time.sleep(1) #

#Entering user/pass into ECMS loggin page
user_elem = driver.find_element_by_name("userid")
pass_elem = driver.find_element_by_name("password")

user_elem.send_keys(my_username)
pass_elem.send_keys(my_password)

submit_button = driver.find_element_by_name("login")
time.sleep(1)
submit_button.click()

#Navigating past the ECMS Alert box and to the planned projects and advertisments portal
WebDriverWait(driver,1).until(EC.alert_is_present(),"I hate waiting")
alert = driver.switch_to_alert()
alert.accept()
time.sleep(1.5)
driver.get("http://www.dot14.state.pa.us/ECMS/SVADVSearch?action=showMainPage")

time.sleep(1.5)
driver.get("http://www.dot14.state.pa.us/ECMS/SVPLPSearch?action=search&LAST_PUBLISHED_DATE_SEARCH_DIRECTION=02&LAST_PUBLISHED_DATE_UNITS=1&LAST_PUBLISHED_DATE_UNIT_CODE=03&SORT_BY_CODE=01&")


planned_html = driver.page_source

soup = BeautifulSoup(planned_html, "html.parser")

 # Let the user actually see something!

#get the table with the project information
tables = soup.findChildren('table')
#results table is the 7th table on the page
table = tables[6]
#table.find_all("tr")[1].find_all('td')

#get a list of table rows which is a list of individual table data
data_rows = table.find_all('tr')[1:]
print(type(data_rows))
project_data = [[td.getText() for td in data_rows[i].findAll('td')]
            for i in range(len(data_rows))]

#print(project_data)

#get a list of links in the table
links = [[a.get('href') for a in table.findChildren("a")]
        for i in range(len(table.findChildren("a")))]
#links is a repeated list because i had it do the search (number of a) times so we just take the first copy of it.
links = links[0]

project_links = []
#sepearates out project links from other links in links. unrelated:itertools#compress
for href in links:
    if "PLP_ID="in href:
        project_links.append(href)


print("length of project links = " + str(len(project_links)))
print("length of project data = " + str(len(project_data)))
print(len(project_links) == len(project_data))

#onto individual planned project page
data_index = 0
for project in project_links[0:]:
    driver.get(url + project)
    project_html = driver.page_source
    soup = BeautifulSoup(project_html, "html.parser")
    PP_tables = soup.findChildren('table')
    publishing_table = PP_tables[6]
    publishing_rows = publishing_table.find_all('tr')[4]
    print(type(publishing_rows))
    planned_rows_data = [[td.getText() for td in publishing_rows.find_all('td')]]
    print(planned_rows_data)

    if "Construction Inspection" in planned_rows_data[0][1]:
        #grabbing the description of work using same names for different table
        publishing_table = PP_tables[14]
        publishing_rows = publishing_table.find_all('tr')[-1]
        planned_rows_data = [[td.getText() for td in publishing_rows.find_all('td')]]
        project_data[data_index].append(planned_rows_data[-1])
        project_data[data_index].append(url + project)
        excel_links.append(project_data[data_index])

    time.sleep(3)
    data_index += 1


driver.quit()
#Write the Data to excel
wb = openpyxl.load_workbook(config.path_to_excel_workbook, keep_vba =True)

wsMain = wb["Main"]

r = wsMain["L1"].value
for array in excel_links:
    #print(str(array) + " is in column " + str(r))
    r +=1
    c = 1
    for element in array:
        if type(element) == list:
            wsMain.cell(row=r, column=c).value = " ".join(element)
        else:
            wsMain.cell(row=r, column=c).value = str(element)

        c +=1
wb.save(config.path_to_excel_workbook)
